import json

from django.db import models
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, StreamingHttpResponse
from django.shortcuts import render
from django import forms
import os
import time

from django.utils.encoding import escape_uri_path
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook

def proceedXLSX(**kwargs):

    output_name = kwargs.get("output_name");
    output_file = kwargs.get("output_file");
    output_sheet_names = kwargs.get("output_sheet_names");

    price_file = kwargs.get("price_file");
    price_sheet_name = kwargs.get("price_sheet_name");

    price_colmn_name = kwargs.get("price_colmn_name");
    output_column_name = kwargs.get("output_column_name");

    price_column_org = kwargs.get("price_column_org");
    price_column_real = kwargs.get("price_column_real");

    output_column_org = kwargs.get("output_column_org");
    output_column_real = kwargs.get("output_column_real");

    dest_tbl = load_workbook (output_file);
    product_tbl = load_workbook (price_file,data_only=True);

    def makePrice(output_sheet_name):
        dest_sheet = dest_tbl [output_sheet_name]
        product_sheet = product_tbl [price_sheet_name];

        product_names = product_sheet [price_colmn_name]; ## 报价
        dest_names = dest_sheet [output_column_name]; ## 账单

        for ceil in dest_names:
            if ceil.value:
                found = False;
                try:
                    for price_ceil in product_names:
                        if price_ceil.value == ceil.value:

                            org_price = product_sheet [price_column_org][price_ceil.row - 1];
                            now_price = product_sheet [price_column_real][price_ceil.row - 1];

                            if org_price and now_price:
                                dest_sheet [output_column_org] [ceil.row - 1].value = round (org_price.value,2);
                                dest_sheet [output_column_real] [ceil.row - 1].value = round (now_price.value,2);
                                found = True;

                            break;
                except Exception as err:
                    pass

                if not found:
                    print ("找不到 %s " % ceil.value)

    for k in range(len(output_sheet_names)):
        output_sheet_name = output_sheet_names [k];
        print("========> 处理 %s " % output_sheet_name)
        makePrice (output_sheet_name);

    output_time = time.strftime("%Y%m%d%H%M%S", time.localtime());
    dest_tbl.save(output_name + "_" + output_time + ".xlsx");

def exportPriceTable(**kwargs):
    try:
        price_file = kwargs.get("price_file");
        price_sheet_name = kwargs.get("price_sheet_name");

        product_tbl = load_workbook(price_file, data_only=True);

        if price_sheet_name == None or price_sheet_name == "":
            price_sheet_name = product_tbl.sheetnames [0];

        price_colmn_name = kwargs.get("price_colmn_name");
        price_column_org = kwargs.get("price_column_org");
        price_column_real = kwargs.get("price_column_real");

        product_sheet = product_tbl[price_sheet_name];
        product_names = product_sheet[price_colmn_name];  ## 报价

        wb = Workbook();
        output_sheet = wb[wb.sheetnames[0]];

        output_sheet.append(["名称", "原始价格", "税后价格"]);

        for price_ceil in product_names:
            org_price = product_sheet[price_column_org][price_ceil.row - 1];
            now_price = product_sheet[price_column_real][price_ceil.row - 1];

            try:
                if org_price and now_price and org_price.value and now_price.value:
                    org_price_value = round(org_price.value, 2);
                    now_price_value = round(now_price.value, 2);
                    output_sheet.append([price_ceil.value, org_price_value, now_price_value]);

            except Exception as err:
                # print (err);
                pass

        output_filename = time.strftime("%Y%m%d%H%M%S_价格表.xlsx", time.localtime());
        output_dir = os.path.join("static","download");
        output_filepath = os.path.join(output_dir, output_filename)

        if not os.path.exists (output_dir):
            os.makedirs(output_dir);

        wb.save(output_filepath);

        return output_filename;

    except Exception as err:
        return "Error " + json.dumps(err.args);


class UploadFileForm(forms.Form):
    title = forms.CharField(max_length=50)
    file = forms.FileField()

def handle_uploaded_file(f):
    with open('name.xlsx', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

def main(request):
    views_name = "Love Django";
    return render(request,"export.html",{"name" : views_name});

def price(request):
    views_name = "Love Django";
    return render(request,"price.html",{"name" : views_name});

@csrf_exempt
def upload(request):
    if request.method == 'POST':
        file_obj = request.FILES.get("file")
        if not file_obj:
            return JsonResponse({
                "ret" : -1,
                "msg" : "No File Uploaded",
            });

        sheet_name  = request.POST["sheet_name"];
        name_column = request.POST["name_column"];
        org_column  = request.POST["org_column"];
        now_column  = request.POST["now_column"];

        resp = exportPriceTable(
            price_file=file_obj,
            price_sheet_name=sheet_name,
            price_colmn_name=name_column,
            price_column_org=org_column,
            price_column_real=now_column,
        )

        if "Error" in resp:
            return JsonResponse({
                "ret" : -1,
                "msg" : resp,
            });

        return JsonResponse({
            "ret" : 0,
            "data" : resp,
            "msg" : "",
        });

def download(request):
    if request.method == 'POST':
        return JsonResponse({
            "ret" : -1,
            "msg" : "Bad Url"
        })

    def file_iterator(file_name, chunk_size=512):
        with open(file_name,'rb') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break

    filename = request.GET ["file"];
    if filename == "" or filename == None:
        return JsonResponse({
            "ret": -1,
            "msg": "Bad Url"
        });

    local_filepath = os.path.join("static","download",filename);

    if not os.path.exists(local_filepath):
        return JsonResponse({
            "ret": -1,
            "msg": "404 file"
        });

    response = StreamingHttpResponse(file_iterator(local_filepath))
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(escape_uri_path(filename))

    return response