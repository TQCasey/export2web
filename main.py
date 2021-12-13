import os
import time

from openpyxl import load_workbook, Workbook


def proceedXLSX(**kwargs):

    output_name = kwargs.get("output_name");
    output_file = kwargs.get("output_file");
    output_sheet_names = kwargs.get("output_sheet_names");

    price_file = kwargs.get("price_file");
    price_sheet_name = kwargs.get("price_sheet_name");

    price_colmn_name = kwargs.get("price_colmn_name");
    output_column_name = kwargs.get("output_column_name");

    price_column_org = kwargs.get("price_column_org");
    price_column_real = kwargs.get("price_column_real");

    output_column_org = kwargs.get("output_column_org");
    output_column_real = kwargs.get("output_column_real");

    dest_tbl = load_workbook (output_file);
    product_tbl = load_workbook (price_file,data_only=True);

    def makePrice(output_sheet_name):
        dest_sheet = dest_tbl [output_sheet_name]
        product_sheet = product_tbl [price_sheet_name];

        product_names = product_sheet [price_colmn_name]; ## 报价
        dest_names = dest_sheet [output_column_name]; ## 账单

        for ceil in dest_names:
            if ceil.value:
                found = False;
                for price_ceil in product_names:
                    if price_ceil.value == ceil.value:

                        org_price = product_sheet [price_column_org][price_ceil.row - 1];
                        now_price = product_sheet [price_column_real][price_ceil.row - 1];

                        if org_price and now_price:
                            # print (price_ceil.row,price_ceil.value,org_price.value,now_price.value,dest_sheet ['G'] [ceil.row - 1].value,dest_sheet ['H'] [ceil.row - 1].value);
                            dest_sheet [output_column_org] [ceil.row - 1].value = round (org_price.value,2);
                            dest_sheet [output_column_real] [ceil.row - 1].value = round (now_price.value,2);
                            found = True;
                        break;

                if not found:
                    print ("找不到 %s " % ceil.value)

    for k in range(len(output_sheet_names)):
        output_sheet_name = output_sheet_names [k];
        print("========> 处理 %s " % output_sheet_name)
        makePrice (output_sheet_name);

    output_time = time.strftime("%Y%m%d%H%M%S", time.localtime());
    dest_tbl.save(output_name + "_" + output_time + ".xlsx");

'''

output_name = "白花社区11月对账单终(1)";
output_file = os.path.join(".",output_name + ".xlsx");
output_sheet_names = ["1","2","3","4","5"];

price_file = os.path.join(".","红岭中学食材新报价(3).xlsx");
price_sheet_name = "综合类（含蔬菜、海鲜水产类、禽蛋类等）";

price_colmn_name = "D";
output_column_name = "B";

price_column_org = "J"; ## 原始价格
price_column_real = "K"; ## 税后价格

output_column_org = "G"; ## 输出原始价格
output_column_real = "H" ; ## 输出实际价格

proceedXLSX (
    output_name = output_name,
    output_file = output_file,
    output_sheet_names = output_sheet_names,

    price_file = price_file,
    price_sheet_name = price_sheet_name,

    price_colmn_name = price_colmn_name,
    output_column_name = output_column_name,

    price_column_org = price_column_org,
    price_column_real = price_column_real,

    output_column_org = output_column_org,
    output_column_real = output_column_real,
)
'''

def exportPriceTable(**kwargs):

    price_file = kwargs.get("price_file");
    price_sheet_name = kwargs.get("price_sheet_name");

    price_colmn_name = kwargs.get("price_colmn_name");
    price_column_org = kwargs.get("price_column_org");
    price_column_real = kwargs.get("price_column_real");

    product_tbl = load_workbook (price_file,data_only=True);
    product_sheet = product_tbl [price_sheet_name];
    product_names = product_sheet [price_colmn_name]; ## 报价

    wb = Workbook ();
    output_sheet = wb [wb.sheetnames [0]];

    output_sheet.append(["名称", "原始价格", "税后价格"]);

    for price_ceil in product_names:
        org_price = product_sheet [price_column_org][price_ceil.row - 1];
        now_price = product_sheet [price_column_real][price_ceil.row - 1];

        try:
            if org_price and now_price and org_price.value and now_price.value:
                org_price_value = round(org_price.value,2);
                now_price_value = round(now_price.value,2);
                output_sheet.append ([price_ceil.value,org_price_value,now_price_value]);

        except Exception as err:
            # print (err);
            pass

    wb.save("价格表.xlsx")

exportPriceTable(
    price_file = "红岭中学食材新报价(3).xlsx",
    price_sheet_name = "综合类（含蔬菜、海鲜水产类、禽蛋类等）",
    price_colmn_name = "D",
    output_column_name = "B",
    price_column_org = "J", ## 原始价格
    price_column_real = "K", ## 税后价格
)
